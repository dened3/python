import cx_Oracle
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, GradientFill
from datetime import datetime

file_name = 'Backlog_.xlsx'
ToDo = 'ToDo'
forPlan = 'Для планирования'
brd = Side(style='thin')
HeadFont = Font(bold=True)

connectJira = 'jira_user/qwe123@jira'

def HeadCellStyle(cell):
	cell.fill   = PatternFill("solid", fgColor="DCDCDC")
	cell.border = Border(left=brd, top=brd, right=brd, bottom=brd)
	cell.font 	= HeadFont

def setHeaderRow(plan_ws, plan_cnt):
	for i in range(ord('A'), ord('K')):
		HeadCellStyle(plan_ws[chr(i) + str(plan_cnt)])

def copy_cell(col, from_cell, to_cell):
	# first value
	to_cell.value = from_cell.value
	# second formatting
	#from_style = ws.cell(row=from_row, column=from_col)._style
	to_cell._style = from_cell._style
	#if from_cell.value is not None:
	if col == 'A':
		to_cell.hyperlink = "https://jira.ftc.ru/browse/" + from_cell.value
    #to_cell.number_format = from_cell.number_format


def makeHeader(ws):
	# Ширина столбцов
	ws.column_dimensions['K'].width = 24 # Новый Статус
	ws.column_dimensions['L'].width = 12 # Затрачено
	# Значения
	ws['K1'].value = 'Статус ' + datetime.today().strftime('%Y-%m-%d')
	ws['L1'].value = 'Затрачено'
	# Bold
	ws['L1'].font = Font(bold=True)
	ws['K1'].font = Font(bold=True)
	# Многострочность
	ws['L1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text = True)
	ws['K1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text = True)
	# Цвет
	ws['L2'].fill = PatternFill("solid", fgColor="DCDCDC")
	ws['K2'].fill = PatternFill("solid", fgColor="DCDCDC")

def create_work_sheet(wb):

	# Выбираем активную вкладку
	plan_ws = wb.create_sheet(forPlan, 1)
	#ws = wb.active
	# Имя вкладки и цвет
	#plan_ws.title = forPlan
	plan_ws.sheet_properties.tabColor = "1072BA"

	plan_ws.column_dimensions['A'].width = 13 # JIRAID
	plan_ws.column_dimensions['B'].width = 80 # Название
	plan_ws.column_dimensions['C'].width = 12 # Срок
	plan_ws.column_dimensions['D'].width = 12 # Приоритет
	plan_ws.column_dimensions['E'].width = 8  # Бизнес
	plan_ws.column_dimensions['F'].width = 12 # Разработка
	plan_ws.column_dimensions['G'].width = 14 # Тестирование
	plan_ws.column_dimensions['H'].width = 14 # Трудозатраты
	plan_ws.column_dimensions['I'].width = 24 # Статус
	plan_ws.column_dimensions['J'].width = 60 # Примечание
	#plan_ws.column_dimensions['K'].width = 22 # Новый Статус

	# Заполняем шапку
	plan_ws['A1'] = 'JIRAID'
	plan_ws['B1'] = 'Название'
	plan_ws['C1'] = 'Срок'
	plan_ws['D1'] = 'Приоритет'
	plan_ws['E1'] = 'Бизнес'
	plan_ws['F1'] = 'Разработка'
	plan_ws['G1'] = 'Тестирование'
	plan_ws['H1'] = 'Трудозатраты(порядок)'
	plan_ws['I1'] = 'Статус'
	plan_ws['J1'] = 'Примечание'

	# Стили
	for col in plan_ws.iter_cols(min_row=1, max_col=10, max_row=1):
		for cell in col:
			cell.font = HeadFont
			cell.border = Border(left=brd, top=brd, right=brd, bottom=brd)
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text = True)

	return plan_ws


if __name__ == "__main__":

	wb = load_workbook(filename = file_name)
	ws = wb[ToDo]

	makeHeader(ws)
	try:
		std = wb.get_sheet_by_name(forPlan)
		wb.remove_sheet(std)
	except Exception as e:
		print('Вкладки "Для планирования" не существует')

	plan_ws = create_work_sheet(wb)

	conJira = cx_Oracle.connect(connectJira)
	cur = conJira.cursor()
	refCurs = conJira.cursor()

	rownum = 0

	plan_cnt = 1

	for c in ws.values:

		rownum += 1

		if rownum != 1:
			stat = ''
			worked = 0

			ws['K' + str(rownum)].border = Border(left=brd, top=brd, right=brd, bottom=brd)
			ws['L' + str(rownum)].border = Border(left=brd, top=brd, right=brd, bottom=brd)

			if c[0] is not None:

				jiraid = str(c[0])
				tprj = jiraid[0:jiraid.find('-')]
				tnum = jiraid[jiraid.find('-')+1:]
				new_jiraid = jiraid

				ssql = """
		select --+ ordered use_nl( p i s )
			   decode(s.pname,
				 'Closed', 'Закрыта',
				 'In Dev', 'В разработке',
				 'In Test', 'В тестировании',
				 'In Analysis', 'Аналитика',
				 'Ready to Dev', 'В очереди на разработку',
				 'Ready to Do', 'В очереди на разработку',
				 'Resolved', 'Готова к тестированию',
				 'Open', 'Открыта',
				 'Delayed', 'Отложена',
				 'Postponed', 'Отложена',
				 'InSupport', 'В сопровождении',
				 s.pname
			   ) status,
			   (
			     select sum(z.worked) worked
  				   from (
				    select round(sum(l.timeworked)/60/60) worked
					  from jira.worklog l
					 where l.issueid = i.id
				    union all
				    select round(sum( l.timeworked / 60 / 60 )) worked
				      from jira.issuelink sl,
					       jira.worklog   l
				     where sl.source = i.id
				       and sl.linktype = 10020
				       and l.issueid = sl.destination
				 ) z
			   ) worked
		  from jira.PROJECT      p,
			   jira.JIRAISSUE    i,
			   jira.issuestatus  s
			where p.pkey = '""" + tprj + """'
			  and i.issuenum = '""" + tnum + """'
			and i.project = p.id
			and s.id = i.issuestatus
		"""

				cur.execute(ssql)
				while True:
					row = cur.fetchone()
					if row is None:
						break
					stat = row[0]
					worked = row[1]
					#print(row[1] + '-->' + row[4])

				if stat == '':
					ssql = """
		select --+ ordered use_nl( m i p s )
			   p.pkey || '-' || i.issuenum   jiraid,
			   decode(s.pname,
					 'Closed', 'Закрыта',
					 'In Dev', 'В разработке',
					 'In Test', 'В тестировании',
					 'In Analysis', 'Аналитика',
					 'Ready to Dev', 'В очереди на разработку',
					 'Ready to Do', 'В очереди на разработку',
					 'Resolved', 'Готова к тестированию',
					 'Open', 'Открыта',
					 'Delayed', 'Отложена',
					 'Postponed', 'Отложена',
					 'InSupport', 'В сопровождении',
					 s.pname
			   ) status,
			   (
				 select sum(z.worked) worked
  				   from (
				    select round(sum(l.timeworked)/60/60) worked
					  from jira.worklog l
					 where l.issueid = i.id
				    union all
				    select round(sum( l.timeworked / 60 / 60 )) worked
				      from jira.issuelink sl,
					       jira.worklog   l
				     where sl.source = i.id
				       and sl.linktype = 10020
				       and l.issueid = sl.destination
				 ) z
			   ) worked
		  from jira.MOVED_ISSUE_KEY m,
			   jira.JIRAISSUE    i,
			   jira.PROJECT      p,
			   jira.issuestatus  s
		  where m.old_issue_key = '""" + jiraid + """'
			and i.id = m.issue_id
			and p.id = i.project
			and s.id = i.issuestatus
		"""
					cur.execute(ssql)
					while True:
						row1 = cur.fetchone()
						if row1 is None:
							break
						else:
							new_jiraid = row1[0]
							stat       = row1[1]
							worked     = row1[2]

				ws['K' + str(rownum)].value = stat
				ws['L' + str(rownum)].value = worked

				#ws['L' + str(rownum)].border = Border(left=brd, top=brd, right=brd, bottom=brd)

				if stat != str(c[8]) and rownum > 1:
					ws['K' + str(rownum)].fill = PatternFill("solid", fgColor="FFFF00")
				#print('K' + str(rownum))
				#print(str(c[0]) + ': ' + str(c[8]) + '-->' + new_jiraid + ': ' + stat)
				print(new_jiraid + ': ' + str(c[8]) + '-->' + stat)
				#'''Для планирования,'''
				if stat not in ('Закрыта', 'InSupport', 'Ready to Support'):
					plan_cnt += 1

					#plan_ws['A' + str(plan_cnt)].value = new_jiraid
					copy_cell('A', ws['A' + str(rownum)], plan_ws['A' + str(plan_cnt)])

					#plan_ws['A' + str(rownum)].style = "Hyperlink"
					plan_ws['A' + str(plan_cnt)].value = new_jiraid
					plan_ws['A' + str(plan_cnt)].hyperlink = "https://jira.ftc.ru/browse/" + new_jiraid

					copy_cell('B', ws['B' + str(rownum)], plan_ws['B' + str(plan_cnt)])
					copy_cell('C', ws['C' + str(rownum)], plan_ws['C' + str(plan_cnt)])
					copy_cell('D', ws['D' + str(rownum)], plan_ws['D' + str(plan_cnt)])
					copy_cell('E', ws['E' + str(rownum)], plan_ws['E' + str(plan_cnt)])
					copy_cell('F', ws['F' + str(rownum)], plan_ws['F' + str(plan_cnt)])
					copy_cell('G', ws['G' + str(rownum)], plan_ws['G' + str(plan_cnt)])
					copy_cell('H', ws['H' + str(rownum)], plan_ws['H' + str(plan_cnt)])
					copy_cell('I', ws['I' + str(rownum)], plan_ws['I' + str(plan_cnt)])
					plan_ws['I' + str(plan_cnt)].value = stat

					#plan_ws['J' + str(plan_cnt)].value = c[9]
					copy_cell('J', ws['J' + str(rownum)], plan_ws['J' + str(plan_cnt)])
			else:
				ws['L' + str(rownum)].fill = PatternFill("solid", fgColor="DCDCDC")
				ws['K' + str(rownum)].fill = PatternFill("solid", fgColor="DCDCDC")
				#
				print(str(rownum) + '-->' + str(plan_cnt))
				plan_cnt += 1
				copy_cell('B', ws['B' + str(rownum)], plan_ws['B' + str(plan_cnt)])
				copy_cell('D', ws['D' + str(rownum)], plan_ws['D' + str(plan_cnt)])
				setHeaderRow(plan_ws, plan_cnt)
				#plan_ws['B' + str(plan_cnt)].value = c[1]
				#plan_ws['B' + str(plan_cnt)].fill = PatternFill("solid", fgColor="FFFF00")



	# Добавляем планирование
	plan_cnt += 1
	plan_ws['B' + str(plan_cnt)].value = 'Для планирования'
	plan_ws['D' + str(plan_cnt)].value = '8000'

	setHeaderRow(plan_ws, plan_cnt)
	#HeadCellStyle(plan_ws['A' + str(plan_cnt)])
	'''
	plan_ws['A' + str(plan_cnt)].fill = PatternFill("solid", fgColor="DCDCDC")
	plan_ws['A' + str(plan_cnt)].border = Border(left=brd, top=brd, right=brd, bottom=brd)
	plan_ws['A' + str(plan_cnt)].font = HeadFont
	plan_ws['B' + str(plan_cnt)].fill = PatternFill("solid", fgColor="DCDCDC")
	plan_ws['B' + str(plan_cnt)].border = Border(left=brd, top=brd, right=brd, bottom=brd)
	plan_ws['B' + str(plan_cnt)].font = HeadFont
	plan_ws['C' + str(plan_cnt)].fill = PatternFill("solid", fgColor="DCDCDC")
	plan_ws['C' + str(plan_cnt)].border = Border(left=brd, top=brd, right=brd, bottom=brd)
	plan_ws['C' + str(plan_cnt)].font = HeadFont
	plan_ws['D' + str(plan_cnt)].fill = PatternFill("solid", fgColor="DCDCDC")
	plan_ws['D' + str(plan_cnt)].border = Border(left=brd, top=brd, right=brd, bottom=brd)
	plan_ws['D' + str(plan_cnt)].font = HeadFont
	plan_ws['E' + str(plan_cnt)].fill = PatternFill("solid", fgColor="DCDCDC")
	plan_ws['E' + str(plan_cnt)].border = Border(left=brd, top=brd, right=brd, bottom=brd)
	plan_ws['E' + str(plan_cnt)].font = HeadFont
	plan_ws['F' + str(plan_cnt)].fill = PatternFill("solid", fgColor="DCDCDC")
	plan_ws['F' + str(plan_cnt)].border = Border(left=brd, top=brd, right=brd, bottom=brd)
	plan_ws['F' + str(plan_cnt)].font = HeadFont
	'''

	#
	psql = """select --+ ordered use_nl( p i s au )
       p.pkey || '-' || i.issuenum                jiraid,
       i.summary                                  summary,
       round( i.timeoriginalestimate / 60 / 60 )  cost,
       s.pname                                    status,
       nvl(au.display_name, i.creator) author,
	   i.creator
  from jira.PROJECT      p,
       jira.JIRAISSUE    i,
       jira.issuestatus  s,
       jira.cwd_user     au
where p.pkey like 'FSGPRJ'
  and i.project = p.id
  and s.id = i.issuestatus
  and s.pname = 'Для планирования'
  --
  and upper(au.user_name(+)) = upper(i.creator)
  and au.directory_id(+) = 10000"""
	#
	cur.execute(psql)
	#
	while True:
		plan_cnt += 1
		row = cur.fetchone()
		if row is None:
			break
		plan_ws['A' + str(plan_cnt)].value = row[0]
		plan_ws['A' + str(plan_cnt)].hyperlink = "https://jira.ftc.ru/browse/" + str(row[0])
		plan_ws['A' + str(plan_cnt)].style = "Hyperlink"
		plan_ws['A' + str(plan_cnt)].border = Border(left=brd, top=brd, right=brd, bottom=brd)
		#
		plan_ws['B' + str(plan_cnt)].value = row[1]
		plan_ws['B' + str(plan_cnt)].border = Border(left=brd, top=brd, right=brd, bottom=brd)
		plan_ws['B' + str(plan_cnt)].alignment = Alignment(horizontal="left", vertical="top", wrap_text = True)

		plan_ws['C' + str(plan_cnt)].border = Border(left=brd, top=brd, right=brd, bottom=brd)

		plan_ws['D' + str(plan_cnt)].value = 0
		plan_ws['D' + str(plan_cnt)].border = Border(left=brd, top=brd, right=brd, bottom=brd)

		plan_ws['E' + str(plan_cnt)].border = Border(left=brd, top=brd, right=brd, bottom=brd)
		if row[5] in ('atrofimova', 'tchesnokova', 'ekponomareva'):
			plan_ws['E' + str(plan_cnt)].value = 'ЗКПК'
		elif row[5] in ('belinsky', 'berezin', 'bilik', 'kochergin'):
			plan_ws['E' + str(plan_cnt)].value = 'ОРР'
		elif row[5] in ('kraus', 'niminuschaja', 'pudul', 'svandreeva', 'ulturgasheva', 'zimina', 'kruglovat'):
			plan_ws['E' + str(plan_cnt)].value = 'ОСП'
		elif row[5] in ('romanenko'):
			plan_ws['E' + str(plan_cnt)].value = 'ДВ'
		else:
			plan_ws['E' + str(plan_cnt)].value = ''

		plan_ws['F' + str(plan_cnt)].border = Border(left=brd, top=brd, right=brd, bottom=brd)
		#plan_ws['F' + str(plan_cnt)].value = row[4]

		plan_ws['G' + str(plan_cnt)].border = Border(left=brd, top=brd, right=brd, bottom=brd)

		plan_ws['H' + str(plan_cnt)].value = row[2]
		plan_ws['H' + str(plan_cnt)].border = Border(left=brd, top=brd, right=brd, bottom=brd)
		plan_ws['I' + str(plan_cnt)].value = row[3]
		plan_ws['I' + str(plan_cnt)].border = Border(left=brd, top=brd, right=brd, bottom=brd)
		plan_ws['J' + str(plan_cnt)].border = Border(left=brd, top=brd, right=brd, bottom=brd)

	#
	cur.close()
	refCurs.close()
	#
	wb.save(file_name)
