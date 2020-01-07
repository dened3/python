#!/usr/bin/python3
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, GradientFill
import re
import os
import datetime
import json

brd = Side(style='thin')

class xlsx:

	def __init__(self, file_name, active_sheet, replace_book=0, replace_sheet=0):
		self.file_name = file_name
		self.sheets = {}
		if replace_book == 1:
			print('rep')
			try:
				os.rename(file_name, file_name[:-4] + datetime.datetime.today().strftime('%Y-%m-%d_%H%M%S') + '.xlsx')
			except Exception as FileNotFoundError:
				z = 1
			self.wb = Workbook()
			self.ws = self.wb.active
			self.sheets[active_sheet] = self.wb.active
			self.ws.title = active_sheet
		else:
			try:
				# Сначала пытаемся найти такой файл
				self.wb = load_workbook(filename = file_name)
				try:
					self.ws = self.wb.get_sheet_by_name(active_sheet)
					self.sheets[active_sheet] = self.wb.get_sheet_by_name(active_sheet)
					if replace_sheet == 1:
						self.wb.remove_sheet(std)
						self.ws = self.wb.create_sheet(active_sheet, 1)
						self.sheets[active_sheet] = self.wb.get_sheet_by_name(active_sheet)
				except Exception as e:
					#print('Вкладки ' + sheet + ' не существует')
					self.ws = self.wb.create_sheet(active_sheet, 1)
					self.sheets[active_sheet] = self.wb.get_sheet_by_name(active_sheet)
			except Exception as e:
				# Если не нашли файл с таким именем, создаем новый
				self.wb = Workbook()
				# Выбираем активную вкладку
				self.ws = self.wb.active
				# меняем название вкладки
				self.ws.title = active_sheet
				self.sheets[active_sheet] = self.wb.get_sheet_by_name(active_sheet)

	def get_sheet(self, sheet_name, replace_sheet=0):
		try:
			self.sheets[sheet_name] = self.wb.get_sheet_by_name(sheet_name)
			if replace_sheet == 1:
				self.wb.remove_sheet(std)
				self.sheets[sheet_name] = self.wb.create_sheet(sheet_name, 1)
		except Exception as e:
			#print('Вкладки ' + sheet + ' не существует')
			self.sheets[sheet_name] = self.wb.create_sheet(sheet_name, 1)

		return self.sheets[sheet_name]

	def save(self):
		self.wb.save(self.file_name)

# Значение и стили для ячейки
def SetCell(cell, value, style):
	style_1 = " " + style.lower() + " "

	if style_1.find(" bold ") != -1:
		cell.font = Font(bold=True)
	if style_1.find(" border ") != -1:
		cell.border = Border(left=brd, top=brd, right=brd, bottom=brd)
	if style_1.find(" center ") != -1:
		cell.alignment = Alignment(horizontal="center", vertical="center")
	if style_1.find(" wrap ") != -1:
		cell.alignment = Alignment(wrap_text = True)
	# Значение
	if value != None:
		cell.value = value
	# Цвет заливки
	if re.search(r' color=(.*?) ', style_1) is not None:
		color = re.search(r'color=(.*?) ', style_1).group(1)
		cell.fill = PatternFill("solid", fgColor=color)

def init_from_smp(smp_file):
	ps = json.loads(open(smp_file).read())
	# Создаем книгу и вкладку
	wb_smp = xlsx(ps['filename'], ps['sheet'], 1)
	ws_smp = wb_smp.get_sheet(ps['sheet'])
	#wb_plan, ws_plan = createWorkBook(ps['filename'], ps['sheet'])
    # Заполняем заголовок
	for i in ps['header']:
		ws_smp.column_dimensions[i].width = ps['header'][i]['width']
		SetCell(ws_smp[i + '1'], ps['header'][i]['name'], "border center bold")
    # Возвращаем книгу, вкладку, имя файла
	return wb_smp


if __name__ == "__main__":

	pl = xlsx('test.xlsx', 'пример', 1)

	aw = pl.ws

	#exit(0)

	aw['A1'] = 'Привет!'
	aw['A2'] = 'Пока!'

	wp = pl.get_sheet('План2')
	wp['A1'] = 'План2!'

	pl.save()
	print(pl.file_name)

	plan_wb = init_from_smp('plan.smp')
	plan_wb.save()
